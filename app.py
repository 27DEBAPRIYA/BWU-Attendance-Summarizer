from flask import Flask, request, send_file, render_template_string
import os
import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

app = Flask(__name__)
UPLOAD_FOLDER = os.path.join(os.getcwd(), 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# HTML Interface
HTML_TEMPLATE = '''
<!DOCTYPE html>
<html>
<head>
    <title>📚BWU Attendance Summarizer</title>
    <style>
        body { font-family: Arial, sans-serif; background-color: #f4f4f4; padding: 30px; text-align: center; }
        h1 { color: #1F4E78; }
        form, .buttons { background: #fff; padding: 20px; border-radius: 8px; display: inline-block; box-shadow: 0 0 10px rgba(0,0,0,0.1); }
        input[type=file] { margin: 10px 0; }
        button, a { background-color: #1F4E78; color: white; padding: 10px 20px; border: none; border-radius: 5px; cursor: pointer; text-decoration: none; }
        button:hover, a:hover { background-color: #163b5c; }
        a.download-link { background: #28a745; }
        a.download-link:hover { background: #218838; }
        h2 { color: #48A6A7; }
        form, .buttons { background: #fff; padding: 20px; border-radius: 8px; display: inline-block; box-shadow: 0 0 10px rgba(0,0,0,0.1); }
        input[type=file] { margin: 10px 0; }
        button, a { background-color: #48A6A7; color: white; padding: 10px 20px; border: none; border-radius: 5px; cursor: pointer; text-decoration: none; }
        button:hover, a:hover { background-color: #163b5c; }
        a.download-link { background: #28a745; }
        a.download-link:hover { background: #218838; }
    </style>
</head>
<body>
    <h1>📊 Format your BWU Attendance File</h1>
    <h2>👇Upload your file to see the Magic👇</h2>

    {% if not excel_ready %}
    <form action="/upload" method="post" enctype="multipart/form-data">
        <label>Select your Excel file (.xlsx only):</label><br>
        <input type="file" name="excel_file" accept=".xlsx" required><br>
        <button type="submit">Upload & Format</button>
    </form>
    {% else %}
    <div class="buttons">
        <a class="download-link" href="/download?filename={{ filename }}">⬇️ Download Formatted File</a><br><br>
        <a href="/" class="btn-secondary">🔁 Upload Another File</a>
    </div>
    {% endif %}
</body>
</html>
'''

# Utility: sanitize filenames
def sanitize_filename(name):
    name = re.sub(r'[^\w\s-]', '', name)
    return re.sub(r'[\s]+', '_', name)

# Utility: check for group/date-like values
def is_date_or_group(value):
    if value is None:
        return False
    val = str(value).lower()
    return (
        "group" in val or
        any(c.isdigit() for c in val) and ('/' in val or '-' in val or ':' in val or val.count("-") >= 2)
    )

@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE, excel_ready=False)

@app.route('/upload', methods=['POST'])
def upload():
    file = request.files['excel_file']
    if not file:
        return 'No file uploaded', 400

    original_name = sanitize_filename(os.path.splitext(file.filename)[0])
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    new_filename = f"Formatted_{original_name}_{timestamp}.xlsx"
    saved_path = os.path.join(UPLOAD_FOLDER, file.filename)
    output_excel_path = os.path.join(UPLOAD_FOLDER, new_filename)

    file.save(saved_path)
    wb = load_workbook(saved_path, data_only=True)
    new_wb = Workbook()
    new_wb.remove(new_wb.active)

    # Styles
    thin = Side(border_style="thin", color="000000")
    thick = Side(border_style="medium", color="000000")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_thick_left = Border(left=thick, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    total_fill = header_fill

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        data = [list(row) for row in ws.iter_rows(values_only=True)]

        # Remove blank columns
        data = list(zip(*data))
        data = [col for col in data if any(cell not in (None, '', ' ') for cell in col)]
        data = list(zip(*data))
        if not data:
            continue

        # Find header row
        header_row_idx = None
        for i, row in enumerate(data):
            if any(str(cell).strip().lower() == "student code" for cell in row if cell):
                header_row_idx = i
                break
        if header_row_idx is None:
            continue

        headers = list(data[header_row_idx])
        student_code_col = next((i for i,h in enumerate(headers) if str(h).strip().lower() == "student code"), None)

        # Determine last row where Student Code starts with "BWU/"
        last_data_row = header_row_idx
        for i in range(header_row_idx+1, len(data)):
            code = str(data[i][student_code_col]).strip() if data[i][student_code_col] else ''
            if code.startswith("BWU/"):
                last_data_row = i

        trimmed_data = data[header_row_idx:last_data_row+1]

        # Remove date/group columns
        valid_cols = [i for i, h in enumerate(trimmed_data[0]) if not is_date_or_group(h)]
        trimmed_data = [[row[i] for i in valid_cols] for row in trimmed_data]
        headers = trimmed_data[0]

        # Create new sheet
        new_ws = new_wb.create_sheet(sheet_name.upper())
        new_ws.page_setup.orientation = new_ws.ORIENTATION_PORTRAIT
        new_ws.page_setup.paperSize = new_ws.PAPERSIZE_A4
        new_ws.print_title_rows = '1:2'

        # Department title
        new_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        dept_cell = new_ws.cell(row=1, column=1, value=sheet_name)
        dept_cell.font = Font(size=14, bold=True)
        dept_cell.alignment = Alignment(horizontal="center", vertical="center")

        present_col = next((i for i,h in enumerate(headers) if "present %" in str(h).lower()), None)
        total_pres_col = next((i for i,h in enumerate(headers) if "total present" in str(h).lower()), None)
        total_class_col = next((i for i,h in enumerate(headers) if "total classes" in str(h).lower()), None)

        # Write rows
        for i, row in enumerate(trimmed_data):
            for j, val in enumerate(row):
                cell = new_ws.cell(row=i+2, column=j+1, value=val)

                # recalc %
                if present_col is not None and j == present_col and i > 0:
                    try:
                        tp = float(row[total_pres_col])
                        tc = float(row[total_class_col])
                        if tc:
                            cell.value = f"{round((tp/tc)*100)}%"
                    except:
                        pass

                if i == 0:
                    cell.fill = header_fill
                    cell.font = white_font
                cell.border = border_thick_left if j == 0 else border_all
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Total row
        total_row_idx = len(trimmed_data) + 2
        label = new_ws.cell(row=total_row_idx, column=1, value="Total")
        label.font = white_font
        label.fill = total_fill
        label.alignment = Alignment(horizontal="center")
        label.border = border_thick_left

        for col in range(2, len(headers)+1):
            header_text = str(new_ws.cell(row=2, column=col).value).lower()
            col_letter = get_column_letter(col)
            cell = new_ws.cell(row=total_row_idx, column=col)

            if "total present" in header_text or "total classes" in header_text:
                cell.value = f"=SUM({col_letter}3:{col_letter}{total_row_idx-1})"
            elif "present %" in header_text:
                tp_letter = get_column_letter(total_pres_col+1)
                tc_letter = get_column_letter(total_class_col+1)
                cell.value = f'=ROUND(({tp_letter}{total_row_idx}/{tc_letter}{total_row_idx})*100,0)&"%"'

            cell.font = white_font
            cell.fill = total_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border_all

        # Auto-fit columns
        for col in range(1, new_ws.max_column+1):
            max_len = max(len(str(c.value)) if c.value is not None else 0 for c in new_ws[get_column_letter(col)])
            new_ws.column_dimensions[get_column_letter(col)].width = min(30, max(10, max_len+2))

    new_wb.save(output_excel_path)
    return render_template_string(HTML_TEMPLATE, excel_ready=True, filename=new_filename)

@app.route('/download')
def download():
    filename = request.args.get("filename")
    file_path = os.path.join(UPLOAD_FOLDER, filename)
    if not os.path.exists(file_path):
        return f"File not found: {file_path}", 404
    return send_file(file_path, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
